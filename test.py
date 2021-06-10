# import openpyxl
# import pandas


# wb = openpyxl.load_workbook('C:\\Users\\Manraj\\Desktop\\python project lib management\\stockDatabase.xlsx')
# df = pandas.read_excel('C:\\Users\\Manraj\\Desktop\\python project lib management\\stockDatabase.xlsx')
# sheet = wb.active


# print(sheet.max_row)
# print(sheet.max_column)

# wb.save('stockDatabase.xlsx')

# import datetime
# import pandas as pd
# import openpyxl

# df = pd.read_excel('booksIssuedDatabase.xlsx')
# df_stock = pd.read_excel('stockDatabase.xlsx')
# wb = openpyxl.load_workbook('booksIssuedDatabase.xlsx')
# sheet = wb.active


# print(df)
# print(df_stock)

import datetime
import openpyxl
import pandas

df = pandas.read_excel('stockDatabase.xlsx')
wb = openpyxl.load_workbook('stockDatabase.xlsx')
sheet = wb.active

isbn = input('book name').lower()

for i in range(len(df['isbn'])):
    if(df['isbn'][i]==isbn):
        print(i)


# t1 = datetime.date.today()
# t2 = datetime.date(2021,06,20)
# print((t2-t1).days)