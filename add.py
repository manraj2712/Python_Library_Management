

def bookDataInput(columnNumber):
    if(columnNumber == 1):
        return input('\nEnter the name of the book\n').lower()
    elif(columnNumber == 2):
        return input('\nEnter the Author for the book\n').lower()
    elif(columnNumber == 3):
        return input('\nEnter the ISBN Code of the book\n').lower()
    elif(columnNumber == 4):
        try:
           return int(input('\nEnter the Quantity Available of the book\n'))
        
        except ValueError:
            return 0

    else:
        print('\nEnter a valid column Number\n')


def addBook():

    import openpyxl as op
    
    wb = op.load_workbook('stockDatabase.xlsx')
    sheet = wb.active

    currentRow = sheet.max_row+1
    currentColumn = 1

    bookName = bookDataInput(1)
    authorName = bookDataInput(2)
    isbn = bookDataInput(3)
    qty = bookDataInput(4)

    while(currentColumn <= 4):
        c = sheet.cell(row=currentRow,column=currentColumn)
        if qty:
            if currentColumn == 1:
                c.value = bookName
            elif currentColumn == 2:
                c.value = authorName
            elif currentColumn == 3:
                c.value = isbn
            elif currentColumn == 4 and qty:
                c.value = qty
        else:
            print('\nEnter a valid quantity in interger!\n')
            break
        currentColumn+=1

    wb.save('stockDatabase.xlsx')
