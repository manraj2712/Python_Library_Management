def returnBookByIsbn():
    import pandas
    import openpyxl
    import datetime

    df = pandas.read_excel('booksIssuedDatabase.xlsx')
    df_stock = pandas.read_excel('stockDatabase.xlsx')
    wb = openpyxl.load_workbook('booksIssuedDatabase.xlsx')
    wb_stock = openpyxl.load_workbook('stockDatabase.xlsx')
    sheet = wb.active
    sheet_stock = wb_stock.active

    studentId = input('\nEnter the Student Id returning the book\n').lower()
    bookIsbn = input('\nEnter the ISBN Code for the book to be returned\n').lower()

    for i in range(len(df['student id'])):

        if(df['student id'][i] == studentId and df['book isbn'][i] == bookIsbn):

            if(not(sheet.cell(row=i+2,column=4).value)):

                returnDate = datetime.datetime.today()
                sheet.cell(row=i+2,column=4).value = returnDate
                issueDate = sheet.cell(row=i+2,column=3).value

                daysDifference = (returnDate-issueDate).days

                if(daysDifference-30>0):
                    sheet.cell(row=i+2,column=5).value = (daysDifference-30)*10
                else:
                    sheet.cell(row=i+2,column=5).value = 0

                for j in range(len(df_stock['isbn'])):
                    if df_stock['isbn'][j] == bookIsbn:
                        sheet_stock.cell(row=j+2,column=4).value = (int(sheet_stock.cell(row=j+2,column=4).value))+1
                        break
                
                print(f'\n Book with ISBN code {bookIsbn} has been successfully Restocked !\n')
    wb.save('booksIssuedDatabase.xlsx')
    wb_stock.save('stockDatabase.xlsx')                


def fineByStudentId():

    import pandas
    import openpyxl

    df = pandas.read_excel('booksIssuedDatabase.xlsx')
    wb = openpyxl.load_workbook('booksIssuedDatabase.xlsx')
    sheet = wb.active

    studentId = input('\nEnter the Student Id to check for fine\n').lower()
    fine = 0
    flag = 0
    for i in range(len(df['student id'])):
        if(df['student id'][i]==studentId):
            try:
                fine += int(sheet.cell(row=i+2,column=5).value)
            except:
                fine = 0
            flag = 1

    if(flag == 0):
        print(f'\nStudent Id {studentId} is not listed with us!\n')
    else:
        print(f'\nFine for Student Id {studentId} is Rupees {fine}\n')

