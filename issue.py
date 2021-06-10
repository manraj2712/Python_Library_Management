
def inputBookIssueData(columnNumber):
    if(columnNumber == 1):
        return input('\nEnter Student Id to which the book is being issued\n').lower()
    elif(columnNumber == 2):
        return input('\nEnter the book isbn being issued\n').lower()
    else:
        print('Enter a valid column Number\n')


def issueBook():
    
    import pandas as pd
    import openpyxl
    import datetime

    df = pd.read_excel('booksIssuedDatabase.xlsx')
    wb = openpyxl.load_workbook('C:\\Users\\Manraj\\Desktop\\python project lib management\\booksIssuedDatabase.xlsx')
    sheet = wb.active
    wb_stock = openpyxl.load_workbook('C:\\Users\\Manraj\\Desktop\\python project lib management\\stockDatabase.xlsx')
    df_stock = pd.read_excel('stockDatabase.xlsx')
    sheet_stock = wb_stock.active


    flag = 0
    currentRow = sheet.max_row+1

    studentId = inputBookIssueData(1)
    isbn = inputBookIssueData(2)


    for i in range(len(df_stock['isbn'])):

        if (df_stock['isbn'][i] == isbn):

            
            print(df_stock.iloc[i])
            print(f'\nDo you want to issue this book to Student id {studentId}\n')
            option = input('''\nEnter 'Y' to confirm and 'N' to Exit\n''').lower()

            if option == 'y':

                currentStock = int(sheet_stock.cell(row=i+2,column=4).value)
                sheet_stock.cell(row=i+2,column=4).value = currentStock-1

                sheet.cell(row=currentRow,column=1).value = studentId
                sheet.cell(row=currentRow,column=2).value = isbn
                sheet.cell(row=currentRow,column=3).value = datetime.date.today()

                print(f'\nBook with ISBN {isbn} has been Issued to Student ID {studentId} for 15 DAYS\n')

                flag = 1
                break

            else:
                flag = 1
                break

    if flag == 0:
        print('\nEntered ISBN Code book is not available in Stock!\n')

    wb.save('booksIssuedDatabase.xlsx')
    wb_stock.save('stockDatabase.xlsx')
